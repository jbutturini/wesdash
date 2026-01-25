from __future__ import annotations

from typing import Any, Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from .formats import normalize_year_column


def write_df(wb: Workbook, sheet_name: str, df: pd.DataFrame, freeze: str = "A2") -> None:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    ws = wb.create_sheet(sheet_name)
    if df.empty:
        ws.append(["empty"])
        return
    df = normalize_year_column(df)
    df = df.where(pd.notnull(df), None)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    ws.freeze_panes = freeze
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for cell in col[:50]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)
    _format_date_columns(ws)


def _format_date_columns(ws) -> None:
    header = [cell.value for cell in ws[1]]
    for idx, col_name in enumerate(header, start=1):
        if col_name in ("year", "period_start"):
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx, max_col=idx):
                cell = row[0]
                if cell.value is not None:
                    cell.number_format = "yyyy-mm-dd"


def build_data_dictionary(schemas: List[Dict[str, Any]]) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for schema in schemas:
        measures = schema.get("measures", {})
        for field, desc in measures.items():
            rows.append({
                "dataset": schema.get("name"),
                "field": field,
                "description": desc,
                "source_name": schema.get("source_name"),
                "source_refresh_cadence": schema.get("source_refresh_cadence"),
                "geo_method": schema.get("geo_method"),
                "limitations": schema.get("limitations"),
            })
    return pd.DataFrame(rows)


def build_workbook(
    output_path: str,
    pipeline_tables: Dict[str, pd.DataFrame],
    households_tables: Dict[str, pd.DataFrame],
    chooser_tables: Dict[str, pd.DataFrame],
    public_alt_df: pd.DataFrame,
    data_dict_df: pd.DataFrame,
) -> None:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    for sheet_name, df in pipeline_tables.items():
        write_df(wb, sheet_name, df)
    for sheet_name, df in households_tables.items():
        write_df(wb, sheet_name, df)
    for sheet_name, df in chooser_tables.items():
        write_df(wb, sheet_name, df)
    write_df(wb, "public_alternatives", public_alt_df)
    write_df(wb, "data_dictionary", data_dict_df)

    wb.save(output_path)
