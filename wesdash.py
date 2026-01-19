#!/usr/bin/env python3
from __future__ import annotations

'''
WES KPI Dashboard Data Puller

Examples:
  # Refresh ACS 5-year + ACS 1-year in one run (2015 through latest published).
  # All output is dropped in the ./out/ folder.
  python wesdash.py refresh --geo geo.yaml

  # Optional: add OSSE chronic absenteeism file as a DC public-alternatives input
  python wesdash.py refresh --geo geo.yaml \
    --osse-chronic-url "https://osse.dc.gov/.../Chronic%20Absenteeism%20Metric%20Scores.xlsx"
'''
import argparse
import os
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import requests
import yaml
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

CENSUS_API_KEY = os.getenv("CENSUS_API_KEY", "").strip()
DEFAULT_OUT_DIR = "out"
DEFAULT_OUT_ACS5 = "wes_kpi_acs5.xlsx"
DEFAULT_OUT_ACS1 = "wes_kpi_acs1.xlsx"
DEFAULT_START_YEAR = 2015


def dataset_exists(year: int, dataset: str) -> bool:
    """Return True if the Census API dataset endpoint exists."""
    # Use variables.json to ensure the dataset is actually published.
    url = f"https://api.census.gov/data/{year}/acs/{dataset}/variables.json"
    try:
        r = requests.get(url, timeout=20)
        return r.status_code == 200
    except requests.RequestException:
        return False


def resolve_latest_year(requested_year: int, dataset: str, max_back: int = 10) -> int:
    """Walk backwards until a dataset vintage exists (prevents confusing 404s)."""
    for y in range(requested_year, requested_year - max_back, -1):
        if dataset_exists(y, dataset):
            return y
    raise RuntimeError(f"Could not find an available {dataset} vintage within the last {max_back} years from {requested_year}.")


def pick_year_for_dataset(dataset: str, year_map: Dict[str, int]) -> int:
    """Map member.dataset (acs5/acs1) to the correct year arg."""
    if dataset.startswith("acs1"):
        return year_map["acs1"]
    if dataset.startswith("acs5"):
        return year_map["acs5"]
    # Default fallback: treat as acs5-like
    return year_map["acs5"]


@dataclass(frozen=True)
class DatasetGeo:
    dataset: str
    for_clause: str
    in_clause: Optional[str] = None


def census_base_url(year: int, dataset: str) -> str:
    return f"https://api.census.gov/data/{year}/acs/{dataset}"


def _normalize_geo_for_year(year: int, geo: DatasetGeo) -> DatasetGeo:
    if year > 2019 and geo.for_clause.startswith("zip code tabulation area:"):
        return DatasetGeo(dataset=geo.dataset, for_clause=geo.for_clause, in_clause=None)
    return geo


def census_get(year: int, dataset: str, variables: List[str], geo: DatasetGeo) -> pd.DataFrame:
    url = census_base_url(year, dataset)
    norm_geo = _normalize_geo_for_year(year, geo)
    params = {"get": ",".join(["NAME"] + variables), "for": norm_geo.for_clause}
    if norm_geo.in_clause:
        params["in"] = norm_geo.in_clause
    if CENSUS_API_KEY:
        params["key"] = CENSUS_API_KEY

    r = requests.get(url, params=params, timeout=60)
    r.raise_for_status()
    data = r.json()
    header, rows = data[0], data[1:]
    df = pd.DataFrame(rows, columns=header)

    for v in variables:
        if v in df.columns:
            df[v] = pd.to_numeric(df[v], errors="coerce")
    return df


def census_get_or_warn(
    year: int,
    dataset: str,
    variables: List[str],
    geo: DatasetGeo,
    geo_key: str,
) -> Optional[pd.DataFrame]:
    try:
        return census_get(year, dataset, variables, geo)
    except requests.HTTPError as exc:
        if exc.response is not None and exc.response.status_code == 400:
            geo_desc = geo.for_clause
            if geo.in_clause:
                geo_desc = f"{geo_desc} in {geo.in_clause}"
            detail = exc.response.text.strip() if exc.response.text else str(exc)
            print(f"Skipping {geo_key} ({dataset} {year}) for {geo_desc}: {detail}")
            return None
        raise


def census_variables_index(year: int, dataset: str, group: Optional[str] = None) -> Dict[str, Any]:
    if group:
        url = f"https://api.census.gov/data/{year}/acs/{dataset}/groups/{group}.json"
    else:
        url = f"https://api.census.gov/data/{year}/acs/{dataset}/variables.json"
    r = requests.get(url, timeout=60)
    r.raise_for_status()
    return r.json()


def census_variables_index_optional(
    year: int,
    dataset: str,
    group: Optional[str],
) -> Optional[Dict[str, Any]]:
    try:
        return census_variables_index(year, dataset, group=group)
    except requests.HTTPError as exc:
        if exc.response is not None and exc.response.status_code == 404:
            return None
        raise


def load_geo_config(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def expand_geos(cfg: Dict[str, Any]) -> Dict[str, List[DatasetGeo]]:
    out: Dict[str, List[DatasetGeo]] = {}
    for key, spec in cfg.get("geographies", {}).items():
        members: List[DatasetGeo] = []
        if "datasets" in spec:
            for ds in spec["datasets"]:
                dataset = ds.get("dataset")
                if not dataset:
                    raise ValueError(f"Geo '{key}' dataset entry is missing 'dataset'")
                if "members" in ds:
                    for member in ds["members"]:
                        members.append(DatasetGeo(dataset=dataset, for_clause=member["for"], in_clause=member.get("in")))
                elif "for" in ds:
                    members.append(DatasetGeo(dataset=dataset, for_clause=ds["for"], in_clause=ds.get("in")))
                else:
                    raise ValueError(f"Geo '{key}' dataset entry must have 'for' or 'members'")
            out[key] = members
        else:
            raise ValueError(f"Geo '{key}' must have 'datasets', 'census', or 'custom'")
    return out


def load_geo_members(path: str) -> Dict[str, List[DatasetGeo]]:
    return expand_geos(load_geo_config(path))


def filter_geo_members(geo_members: Dict[str, List[DatasetGeo]], dataset_prefix: str) -> Dict[str, List[DatasetGeo]]:
    filtered: Dict[str, List[DatasetGeo]] = {}
    for key, members in geo_members.items():
        keep = [m for m in members if m.dataset.startswith(dataset_prefix)]
        if keep:
            filtered[key] = keep
    return filtered


# -----------------------------
# KPI pulls
# -----------------------------
B01001_VARS = {
    "age0_4": ["B01001_003E", "B01001_027E"],
    "age5_9": ["B01001_004E", "B01001_028E"],
    "age10_14": ["B01001_005E", "B01001_029E"],
}

S1101_VAR_HH_OWN_CHILDREN_U18 = "S1101_C01_005E"
B19131_GROUP = "B19131"

B14003_VARS = {
    "pub_3_4": ["B14003_004E", "B14003_032E"],
    "pub_5_9": ["B14003_005E", "B14003_033E"],
    "pub_10_14": ["B14003_006E", "B14003_034E"],
    "priv_3_4": ["B14003_013E", "B14003_041E"],
    "priv_5_9": ["B14003_014E", "B14003_042E"],
    "priv_10_14": ["B14003_015E", "B14003_043E"],
}


def pull_pipeline_acs(
    acs5_year: int,
    acs1_year: int,
    geo_cfg_path: str,
    geo_members: Optional[Dict[str, List[DatasetGeo]]] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    geo_members = geo_members or load_geo_members(geo_cfg_path)
    raw_rows, out_rows = [], []

    for geo_key, members in geo_members.items():
        agg = {"geo_key": geo_key, "age0_4": 0.0, "age5_9": 0.0, "age10_14": 0.0}
        year_val: Optional[int] = None
        geo_raw_rows: List[Dict[str, Any]] = []
        skip_geo = False
        vars_needed = sorted({v for vs in B01001_VARS.values() for v in vs})
        for m in members:
            y = acs5_year if m.dataset == "acs5" else (acs1_year if acs1_year is not None else acs5_year)
            if year_val is None:
                year_val = y
            df = census_get_or_warn(y, m.dataset, vars_needed, DatasetGeo(m.dataset, m.for_clause, m.in_clause), geo_key)
            if df is None:
                skip_geo = True
                break
            row = df.iloc[0].to_dict()
            row.update({
                "geo_key": geo_key,
                "member_for": m.for_clause,
                "member_in": m.in_clause or "",
                "dataset": m.dataset,
                "year": y,
            })
            geo_raw_rows.append(row)

            agg["age0_4"] += float(df[B01001_VARS["age0_4"]].sum(axis=1).iloc[0])
            agg["age5_9"] += float(df[B01001_VARS["age5_9"]].sum(axis=1).iloc[0])
            agg["age10_14"] += float(df[B01001_VARS["age10_14"]].sum(axis=1).iloc[0])

        if skip_geo:
            continue
        raw_rows.extend(geo_raw_rows)
        agg["year"] = year_val
        out_rows.append(agg)

    return pd.DataFrame(raw_rows), pd.DataFrame(out_rows)


def pull_households_acs(
    acs5_year: int,
    acs1_year: int,
    geo_cfg_path: str,
    subject_dataset: str = "auto",
    geo_members: Optional[Dict[str, List[DatasetGeo]]] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    geo_members = geo_members or load_geo_members(geo_cfg_path)
    raw_rows, out_rows = [], []

    for geo_key, members in geo_members.items():
        agg_val = 0.0
        year_val: Optional[int] = None
        geo_raw_rows: List[Dict[str, Any]] = []
        skip_geo = False
        for m in members:
            # Choose subject endpoint based on the member dataset unless explicitly overridden.
            subj = subject_dataset
            if subject_dataset == "auto":
                subj = "acs5/subject" if m.dataset == "acs5" else "acs1/subject"
            y = acs5_year if m.dataset == "acs5" else (acs1_year if acs1_year is not None else acs5_year)
            if year_val is None:
                year_val = y
            df = census_get_or_warn(y, subj, [S1101_VAR_HH_OWN_CHILDREN_U18], DatasetGeo(subj, m.for_clause, m.in_clause), geo_key)
            if df is None:
                skip_geo = True
                break
            row = df.iloc[0].to_dict()
            row.update({
                "geo_key": geo_key,
                "member_for": m.for_clause,
                "member_in": m.in_clause or "",
                "dataset": subj,
                "year": y,
            })
            geo_raw_rows.append(row)
            agg_val += float(df[S1101_VAR_HH_OWN_CHILDREN_U18].iloc[0])
        if skip_geo:
            continue
        raw_rows.extend(geo_raw_rows)
        out_rows.append({"geo_key": geo_key, "hh_own_children_u18": agg_val, "year": year_val})

    return pd.DataFrame(raw_rows), pd.DataFrame(out_rows)


def _select_b19131_income_vars(meta: Dict[str, Any], income_labels: List[str]) -> List[str]:
    variables = meta.get("variables", {})
    selected: List[str] = []
    for var, info in variables.items():
        if not var.endswith("E"):
            continue
        label = info.get("label", "")
        if "With own children of the householder under 18 years" not in label:
            continue
        if any(il in label for il in income_labels):
            selected.append(var)
    return sorted(set(selected))


def pull_high_income_acs(
    acs5_year: int,
    acs1_year: int,
    geo_cfg_path: str,
    geo_members: Optional[Dict[str, List[DatasetGeo]]] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    geo_members = geo_members or load_geo_members(geo_cfg_path)

    # Variable IDs can vary slightly by vintage. Discover variables by LABEL.
    datasets_present = {m.dataset for ms in geo_members.values() for m in ms}
    year_map = {"acs5": acs5_year, "acs1": acs1_year}

    meta_cache: Dict[str, Dict[str, Any]] = {}
    for ds in datasets_present:
        meta = census_variables_index_optional(year_map[ds], ds, group=B19131_GROUP)
        if meta is None:
            print(f"Skipping high-income for {ds} {year_map[ds]}: B19131 not available.")
            return pd.DataFrame(), pd.DataFrame()
        meta_cache[ds] = meta

    vars_cache: Dict[str, Dict[str, List[str]]] = {}
    for ds, meta in meta_cache.items():
        v150 = _select_b19131_income_vars(meta, ["$150,000 to $199,999"])
        v200 = _select_b19131_income_vars(meta, ["$200,000 or more"])
        allv = sorted(set(v150 + v200))
        if not allv:
            raise RuntimeError(f"Could not discover B19131 high-income variables for {ds} in the requested vintage.")
        vars_cache[ds] = {"150_199": v150, "200_plus": v200, "all": allv}

    raw_rows, out_rows = [], []

    for geo_key, members in geo_members.items():
        agg_150, agg_200 = 0.0, 0.0
        year_val: Optional[int] = None
        geo_raw_rows: List[Dict[str, Any]] = []
        skip_geo = False
        for m in members:
            dataset = m.dataset
            y = acs5_year if dataset == "acs5" else (acs1_year if acs1_year is not None else acs5_year)
            if year_val is None:
                year_val = y
            v = vars_cache[dataset]["all"]
            df = census_get_or_warn(y, dataset, v, DatasetGeo(dataset, m.for_clause, m.in_clause), geo_key)
            if df is None:
                skip_geo = True
                break
            row = df.iloc[0].to_dict()
            row.update({
                "geo_key": geo_key,
                "member_for": m.for_clause,
                "member_in": m.in_clause or "",
                "dataset": dataset,
                "year": y,
            })
            geo_raw_rows.append(row)

            v150 = vars_cache[dataset]["150_199"]
            v200 = vars_cache[dataset]["200_plus"]
            if v150:
                agg_150 += float(df[v150].sum(axis=1).iloc[0])
            if v200:
                agg_200 += float(df[v200].sum(axis=1).iloc[0])

        if skip_geo:
            continue
        raw_rows.extend(geo_raw_rows)
        out_rows.append({
            "geo_key": geo_key,
            "hhkids_income_150_199": agg_150,
            "hhkids_income_200_plus": agg_200,
            "hhkids_income_150_plus": agg_150 + agg_200,
            "year": year_val,
        })

    return pd.DataFrame(raw_rows), pd.DataFrame(out_rows)


def pull_chooser_rate_acs(
    acs5_year: int,
    acs1_year: int,
    geo_cfg_path: str,
    geo_members: Optional[Dict[str, List[DatasetGeo]]] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    geo_members = geo_members or load_geo_members(geo_cfg_path)
    vars_needed = sorted({v for vs in B14003_VARS.values() for v in vs})

    raw_rows, out_rows = [], []
    for geo_key, members in geo_members.items():
        pub, priv = 0.0, 0.0
        year_val: Optional[int] = None
        geo_raw_rows: List[Dict[str, Any]] = []
        skip_geo = False
        for m in members:
            dataset = m.dataset
            y = acs5_year if dataset == "acs5" else (acs1_year if acs1_year is not None else acs5_year)
            if year_val is None:
                year_val = y
            df = census_get_or_warn(y, dataset, vars_needed, DatasetGeo(dataset, m.for_clause, m.in_clause), geo_key)
            if df is None:
                skip_geo = True
                break
            row = df.iloc[0].to_dict()
            row.update({
                "geo_key": geo_key,
                "member_for": m.for_clause,
                "member_in": m.in_clause or "",
                "dataset": dataset,
                "year": y,
            })
            geo_raw_rows.append(row)

            pub += float(df[B14003_VARS["pub_3_4"]].sum(axis=1).iloc[0]) + float(df[B14003_VARS["pub_5_9"]].sum(axis=1).iloc[0]) + float(df[B14003_VARS["pub_10_14"]].sum(axis=1).iloc[0])
            priv += float(df[B14003_VARS["priv_3_4"]].sum(axis=1).iloc[0]) + float(df[B14003_VARS["priv_5_9"]].sum(axis=1).iloc[0]) + float(df[B14003_VARS["priv_10_14"]].sum(axis=1).iloc[0])

        if skip_geo:
            continue
        raw_rows.extend(geo_raw_rows)
        chooser = (priv / (priv + pub)) if (priv + pub) > 0 else float("nan")
        out_rows.append({
            "geo_key": geo_key,
            "public_enrolled_3_14": pub,
            "private_enrolled_3_14": priv,
            "private_chooser_rate_3_14": chooser,
            "year": year_val,
        })

    return pd.DataFrame(raw_rows), pd.DataFrame(out_rows)


# -----------------------------
# DC public alternatives component (OSSE chronic absenteeism xlsx)
# -----------------------------
def pull_osse_chronic_absenteeism(url: str) -> pd.DataFrame:
    import io
    r = requests.get(url, timeout=120)
    r.raise_for_status()
    bio = io.BytesIO(r.content)
    df = pd.read_excel(bio)
    df["source_url"] = url
    df["pulled_at_utc"] = datetime.utcnow().isoformat()
    return df


# -----------------------------
# Excel helpers
# -----------------------------
def ensure_workbook(path: str) -> Workbook:
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    return wb


def write_df(wb: Workbook, sheet_name: str, df: pd.DataFrame, freeze: str = "A2") -> None:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    ws = wb.create_sheet(sheet_name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    ws.freeze_panes = freeze
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for cell in col[:50]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)


def build_kpi_calcs(pipeline: pd.DataFrame, households: pd.DataFrame, high_income: pd.DataFrame, chooser: pd.DataFrame) -> pd.DataFrame:
    df = pipeline.merge(households, on=["geo_key", "year"], how="left") \
                 .merge(high_income, on=["geo_key", "year"], how="left") \
                 .merge(chooser, on=["geo_key", "year"], how="left")

    return df.rename(columns={
        "age0_4": "Age 0-4 count",
        "age5_9": "Age 5-9 count",
        "age10_14": "Age 10-14 count",
        "hh_own_children_u18": "HH w/ own children <18",
        "hhkids_income_150_plus": "High-income HH w/kids (>=150k)",
        "hhkids_income_200_plus": "High-income HH w/kids (>=200k)",
        "private_chooser_rate_3_14": "Private school chooser rate (ages 3-14)",
    })


def ensure_out_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def available_years(start_year: int, end_year: int, dataset: str, strict: bool) -> List[int]:
    years: List[int] = []
    for y in range(start_year, end_year + 1):
        if dataset_exists(y, dataset):
            years.append(y)
            continue
        msg = f"{dataset} {y} dataset not available."
        if strict:
            raise RuntimeError(msg)
        print(f"Skipping {msg}")
    return years


def concat_frames(frames: List[pd.DataFrame]) -> pd.DataFrame:
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


def collect_refresh_time_series(
    args: argparse.Namespace,
    geo_members: Dict[str, List[DatasetGeo]],
    years: List[int],
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    raw_p_list: List[pd.DataFrame] = []
    raw_h_list: List[pd.DataFrame] = []
    raw_i_list: List[pd.DataFrame] = []
    raw_c_list: List[pd.DataFrame] = []
    kpi_p_list: List[pd.DataFrame] = []
    kpi_h_list: List[pd.DataFrame] = []
    kpi_i_list: List[pd.DataFrame] = []
    kpi_c_list: List[pd.DataFrame] = []

    for year in years:
        raw_p, kpi_p = pull_pipeline_acs(year, year, args.geo, geo_members=geo_members)
        raw_h, kpi_h = pull_households_acs(year, year, args.geo, args.subject_dataset, geo_members=geo_members)
        raw_i, kpi_i = pull_high_income_acs(year, year, args.geo, geo_members=geo_members)
        raw_c, kpi_c = pull_chooser_rate_acs(year, year, args.geo, geo_members=geo_members)

        raw_p_list.append(raw_p)
        raw_h_list.append(raw_h)
        raw_i_list.append(raw_i)
        raw_c_list.append(raw_c)
        kpi_p_list.append(kpi_p)
        kpi_h_list.append(kpi_h)
        kpi_i_list.append(kpi_i)
        kpi_c_list.append(kpi_c)

    return (
        concat_frames(raw_p_list),
        concat_frames(raw_h_list),
        concat_frames(raw_i_list),
        concat_frames(raw_c_list),
        concat_frames(kpi_p_list),
        concat_frames(kpi_h_list),
        concat_frames(kpi_i_list),
        concat_frames(kpi_c_list),
    )


def write_refresh_workbook(
    out_path: str,
    raw_p: pd.DataFrame,
    raw_h: pd.DataFrame,
    raw_i: pd.DataFrame,
    raw_c: pd.DataFrame,
    kpi_p: pd.DataFrame,
    kpi_h: pd.DataFrame,
    kpi_i: pd.DataFrame,
    kpi_c: pd.DataFrame,
    osse_df: Optional[pd.DataFrame] = None,
) -> None:
    wb = ensure_workbook(out_path)

    write_df(wb, "RAW_ACS_B01001", raw_p)
    write_df(wb, "RAW_ACS_S1101", raw_h)
    write_df(wb, "RAW_ACS_B19131", raw_i)
    write_df(wb, "RAW_ACS_B14003", raw_c)

    write_df(wb, "KPI_Pipeline", kpi_p)
    write_df(wb, "KPI_Households", kpi_h)
    write_df(wb, "KPI_HighIncome", kpi_i)
    write_df(wb, "KPI_Chooser", kpi_c)

    calcs = build_kpi_calcs(kpi_p, kpi_h, kpi_i, kpi_c)
    write_df(wb, "KPI_Calcs", calcs)

    if osse_df is not None:
        write_df(wb, "RAW_OSSE_ChronicAbs", osse_df)

    wb.save(out_path)
    print(f"Refreshed {out_path}")


# -----------------------------
# CLI
# -----------------------------
def cmd_refresh(args: argparse.Namespace) -> None:
    # Always use the latest published vintages for both datasets.
    current_year = datetime.utcnow().year
    latest_acs5 = resolve_latest_year(current_year, "acs5")
    latest_acs1 = resolve_latest_year(current_year, "acs1")

    geo_members = load_geo_members(args.geo)
    ensure_out_dir(DEFAULT_OUT_DIR)

    osse_df = pull_osse_chronic_absenteeism(args.osse_chronic_url) if args.osse_chronic_url else None

    geo_members_acs5 = filter_geo_members(geo_members, "acs5")
    geo_members_acs1 = filter_geo_members(geo_members, "acs1")

    out_acs5 = os.path.join(DEFAULT_OUT_DIR, DEFAULT_OUT_ACS5)
    if geo_members_acs5:
        years_acs5 = available_years(DEFAULT_START_YEAR, latest_acs5, "acs5", strict=False)
        if years_acs5:
            frames_acs5 = collect_refresh_time_series(args, geo_members_acs5, years_acs5)
            write_refresh_workbook(out_acs5, *frames_acs5, osse_df=osse_df)
        else:
            print(f"No ACS5 years available from {DEFAULT_START_YEAR} to {latest_acs5}; skipped {out_acs5}")
    else:
        print(f"No ACS5 geographies in {args.geo}; skipped {out_acs5}")

    out_acs1 = os.path.join(DEFAULT_OUT_DIR, DEFAULT_OUT_ACS1)
    if geo_members_acs1:
        years_acs1 = available_years(DEFAULT_START_YEAR, latest_acs1, "acs1", strict=False)
        frames_acs1 = collect_refresh_time_series(args, geo_members_acs1, years_acs1)
        write_refresh_workbook(out_acs1, *frames_acs1, osse_df=osse_df)
    else:
        print(f"No ACS1 geographies in {args.geo}; skipped {out_acs1}")


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="wesdash", description="Pull WES KPI datasets into an Excel workbook.")
    sub = p.add_subparsers(dest="cmd", required=True)

    refresh = sub.add_parser("refresh", help="Refresh everything and rebuild KPI_Calcs.")
    refresh.add_argument("--geo", type=str, default="geo.yaml")
    refresh.add_argument("--subject-dataset", type=str, default="auto", help="Subject dataset endpoint: auto, acs5/subject, or acs1/subject")
    refresh.add_argument("--osse-chronic-url", type=str, default="")
    refresh.set_defaults(func=cmd_refresh)

    return p


def main(argv: Optional[List[str]] = None) -> None:
    args = build_parser().parse_args(argv)
    args.func(args)


if __name__ == "__main__":
    main()
